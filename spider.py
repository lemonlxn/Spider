# /usr/bin/env python
# -*- coding:utf-8 -*-

# @Time    : 2018/7/5 21:28
# @Author  : lemon
import json
import re
import time
from datetime import datetime
import cymysql

from openpyxl import Workbook
from openpyxl.styles import Font, colors
import requests

from lxml import html


class Spider:
    def __init__(self):
        self.book_list = []
        self.wb = Workbook()
        self.ws = self.wb.active
        self.font = Font(sz=13, color=colors.RED)


    def save(self, keyword,page=1):

        print('现在开始爬取亚马逊')
        self.__amazon(keyword,page)
        time.sleep(1)

        print('现在开始爬取当当网')
        self.__dangdang(keyword,page)
        time.sleep(1)

        print('现在开始爬取京东网')
        self.__jd(keyword,page)
        time.sleep(1)

        print('现在开始爬取一号店')
        self.__one(keyword,page)
        time.sleep(1)

        print('现在开始爬取淘宝网')
        self.__taobao(keyword,page)
        time.sleep(1)


        print('---------优化数据--------')

        self.book_list = self.__improve_data(self.book_list)

        print('----打印抓取的数据，并排序-----')

        self.book_list = sorted(self.book_list, key=lambda x:float(x['price']), reverse=False)

        for book in self.book_list:
            print(book)

        print('----现在开始保存数据----')

        self.__workbook(keyword, page,self.book_list)

        for book in self.book_list:

            self.__save_to_mysql(book)
            self.ws.append([book['title'], book['price'], book['link'], book['store'], book['data_from'], book['up_time']])

        self.wb.save(r'../static/{}.xlsx'.format(keyword))



    def __amazon(self, keyword,page=1):
        for num in range(int(page)):
            url = 'https://www.amazon.cn/s/ref=sr_pg_{0}?keywords={1}'.format(num+1,keyword)
            r = requests.get(url)

            s = html.fromstring(r.text)

            data = s.xpath('//div[@id="resultsCol"]//div[@id="atfResults"]/ul[@id="s-results-list-atf"]/li')

            for i in data:
                title = i.xpath('div[@class="s-item-container"]//div[@class="a-row a-spacing-none"]/a/@title')
                price = i.xpath('div[@class="s-item-container"]//a/span[@class="a-size-base a-color-price s-price a-text-bold"]/text()')
                link =  i.xpath('div[@class="s-item-container"]//div[@class="a-row a-spacing-none"]/a/@href')
                store = ''

                self.book_list.append({
                    'title': title[0] if title else '',
                    'price': str(price[0]).replace('￥', '').replace(',','') if price else 0,
                    'link': link[0] if link else '',
                    'store': store[0] if store else '',
                    'data_from': '亚马逊',
                    'up_time': str(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                })



    def __dangdang(self, keyword,page=1):
        for num in range(int(page)):
            url = 'http://search.dangdang.com/'

            r = requests.get(url, params={
                'key': keyword,
                'act': 'input',
                'page_index':num+1
            })

            s = html.fromstring(r.text)

            data = s.xpath('//div[@id="search_nature_rg"]/ul/li')

            for i in data:
                title = i.xpath('a/@title')
                price = i.xpath('div[@class="ebook_buy"]/p[@class="price e_price"]/span[@class="search_now_price"]/text()') \
                        or i.xpath('p[@class="price"]/span/text()')
                link = i.xpath('a/@href')
                store = i.xpath('p[@class="search_shangjia"]/a/@title')

                self.book_list.append({
                    'title': title[0] if title else '',
                    'price': str(price[0]).replace('¥', '').replace(',','') if price else 0,
                    'link': link[0] if link else '',
                    'store': '当当自营' if store == [] else store[0],
                    'data_from': '当当网',
                    'up_time': str(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                })

    def __jd(self, keyword,page=1):
        for num in range(int(page)):
            url = 'https://search.jd.com/Search?keyword={0}&enc=utf-8&page={1}'.format(keyword,2*num + 1)
            r = requests.get(url)

            r.encoding = 'utf-8'
            s = html.fromstring(r.text)

            data = s.xpath('//div[@id="J_goodsList"]/ul/li')

            for i in data:
                title = i.xpath('div//div[@class="p-name"]/a/@title') or i.xpath('div//div[@class="p-name p-name-type-2"]/a/@title')
                price = i.xpath('div//div[@class="p-price"]/strong/i/text()')
                link = i.xpath('div//div[@class="p-name"]/a/@href') or i.xpath('div//div[@class="p-name p-name-type-2"]/a/@href')
                store = i.xpath('div//div[@class="p-shopnum"]/a/@title') or i.xpath('div//div[@class="p-shop"]/span/a/@title')

                self.book_list.append({
                    'title': title[0] if title else '',
                    'price': str(price[0]).replace(',','') if price else 0,
                    'link': 'https:' + link[0] if link else '',
                    'store': store[0] if store else '',
                    'data_from': '京东',
                    'up_time': str(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                })



    def __one(self, keyword,page=1):
        for num in range(int(page)):
            url = 'http://search.yhd.com/c0-0/k{0}/#page={1}&sort=1'.format(keyword,num+1)
            r = requests.get(url)
            s = html.fromstring(r.text)

            data = s.xpath('//div[@id="itemSearchList"]/div')

            for i in data:
                title = i.xpath('div//p[@class="proName clearfix"]/a/@title')
                price = i.xpath('div//p[@class="proPrice"]/em/@yhdprice')
                link = i.xpath('div//p[@class="proName clearfix"]/a/@href')
                store = i.xpath('div//p[@class="storeName limit_width"]/a/@title')

                self.book_list.append({
                    'title': title[0] if title else '',
                    'price': str(price[0]).replace('¥', '').replace(',','') if price else 0,
                    'link': 'https:' + link[0] if link else '',
                    'store': store[0] if store else '',
                    'data_from': '一号店',
                    'up_time': str(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                })


    def __taobao(self, keyword,page=1):
        try:
            for num in range(int(page)):
                url = 'https://s.taobao.com/search?q={0}&s={1}&ie=utf8'.format(keyword, (num) * 44)
                r = requests.get(url)
                s = r.text

                data = re.findall(r'g_page_config = (.*?) g_srp_loadCss', s, re.S)[0].strip()[:-1]

                data = json.loads(data)

                data_list = data['mods']['itemlist']['data']['auctions']

                for item in data_list:
                    title = item['raw_title']
                    price = item['view_price']
                    link = item['detail_url']
                    store = item['nick']

                    self.book_list.append({
                        'title': title if title else '',
                        'price': str(price).replace(',','') if price else 0,
                        'link': 'https:' + link if link else '',
                        'store': store if store else '',
                        'data_from': '淘宝',
                        'up_time': str(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                    })
        except:
            pass


    def __workbook(self,keyword,page,data=[]):
        self.ws.title = '抓取回来的数据'
        self.ws['A1'] = '您输入的关键词是：' + str(keyword)
        self.ws['A2'] = '您抓取的页数为：'+ str(page)
        self.ws['A1'].font = self.font
        self.ws['A2'].font = self.font
        self.ws['A3'] = '现在抓取的时间是：' + str(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        self.ws['A4'] = '本次抓取的数据量为：'+str(len(data))+'条'
        self.ws['A5'] = ''

        self.ws.append(['标题', '价格', '链接', '店铺', '来源','更新时间'])

    @staticmethod
    def _str_or_num(price_data):
        if price_data.strip().isalpha():
            return True
        else:
            return False


    def __improve_data(self,spider_old_list):
        '''
        去重爬取的数据,以及去除商品价格中含有 '免费'，'不收费'字段等
        :param spider_old_list: 接收爬取的数据
        :return: 返回去重后,且将价格中 '免费'，'不收费'字段等统一设置为 '0' 的数据。
        '''
        spider_list     = []
        spider_new_list = []

        for i in spider_old_list:
            if i not in spider_list:
                spider_list.append(i)


        for items in spider_list:
            if self._str_or_num(str(items['price'])):
                items['price'] = '0'
            spider_new_list.append(items)

        return spider_new_list


    def __save_to_mysql(self,data):
        db = cymysql.connect(
            user = 'root',
            port = 3306,
            passwd = 'woailxn',
            host = 'localhost',
            db   = 'search',
            charset = 'utf8'
        )
        cursor = db.cursor()

        try:
            sql = """INSERT INTO SAVE_DATA(TITLE,PRICE,LINK,STORE,DATA_FROM,UP_TIME) VALUES ('%s','%s','%s','%s','%s','%s');"""\
                  %(data['title'],data['price'],data['link'],data['store'],data['data_from'],data['up_time'])

            cursor.execute(sql)
            cursor.close()
            db.commit()

        except Exception as err:
            db.rollback()
            raise err

        finally:
            db.close()


class Spider_keyword(Spider):

    pass


